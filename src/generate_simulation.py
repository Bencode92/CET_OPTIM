#!/usr/bin/env python3
"""CET RGDU 2026 Simulation Generator

Usage:
    python generate_simulation.py <input_file.xlsx|csv> [output_file.xlsx]

Reads a Marges file and produces an Excel workbook with:
- Paramètres sheet (RGDU 2026 parameters, editable)
- Données sheet (raw data)
- Simulation sheet (6 CET scenarios with Excel formulas)
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys, os

INPUT_FILE = sys.argv[1] if len(sys.argv) > 1 else "data/template_marges.csv"
OUTPUT_FILE = sys.argv[2] if len(sys.argv) > 2 else "output/simulation_cet_rgdu2026.xlsx"
os.makedirs(os.path.dirname(OUTPUT_FILE) or '.', exist_ok=True)

ext = os.path.splitext(INPUT_FILE)[1].lower()
df = pd.read_csv(INPUT_FILE) if ext == '.csv' else pd.read_excel(INPUT_FILE)

col_map = {}
for c in df.columns:
    cl = c.strip().lower()
    if 'hrs' in cl and 'paye' in cl: col_map['hrs'] = c
    elif 'brut' in cl and 'réel' in cl: col_map['brut'] = c
    elif 'ifm' in cl and 'réel' in cl: col_map['ifm'] = c
    elif 'iccp' in cl and 'réel' in cl: col_map['iccp'] = c
    elif 'réduction' in cl: col_map['reduction'] = c
    elif 'cot' in cl and 'pat' in cl: col_map['cotpat'] = c
    elif 'marge' == cl.strip() or ('marge' in cl and '%' not in cl): col_map['marge'] = c
    elif 'matricule' in cl: col_map['matricule'] = c

required = ['hrs', 'brut', 'ifm', 'iccp']
missing = [k for k in required if k not in col_map]
if missing:
    print(f"ERROR: Missing columns: {missing}")
    sys.exit(1)

df_clean = df.dropna(subset=[col_map['hrs']])
df_clean = df_clean[df_clean[col_map['ifm']] > 0].copy()
n_rows = len(df_clean)

# Styles
bold_white = Font(name='Arial', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
param_fill = PatternFill('solid', fgColor='FFFF00')
blue_font = Font(name='Arial', color='0000FF', size=10)
green_font = Font(name='Arial', color='008000', size=10)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

wb = Workbook()

# PARAMÈTRES
ws_p = wb.active
ws_p.title = "Paramètres"
params = [
    ["PARAMÈTRES RGDU 2026", "", ""],
    ["", "", ""],
    ["Paramètre", "Valeur", "Explication"],
    ["SMIC horaire 2026", 12.02, "SMIC brut horaire au 01/01/2026"],
    ["Seuil d'extinction", 3, "Réduction s'annule à 3x SMIC"],
    ["Tmin", 0.02, "Taux minimum (2%)"],
    ["Tmax (FNAL 0,50%)", 0.4013, "Taux max pour FNAL 0,50%"],
    ["Tmax (FNAL 0,10%)", 0.3973, "Taux max pour FNAL 0,10%"],
    ["Tδ (FNAL 0,50%)", 0.3813, "Tmax - Tmin"],
    ["Tδ (FNAL 0,10%)", 0.3773, "Tmax - Tmin"],
    ["Puissance", 1.75, "Exposant courbe"],
    ["", "", ""],
    ["VOTRE ENTREPRISE", "", ""],
    ["FNAL applicable", 0.50, "MODIFIABLE: 0.10 ou 0.50"],
    ["", "", ""],
    ["CALCULÉS", "", ""],
    ["Tmax utilisé", "=IF(B14=0.5,B7,B8)", "Auto selon FNAL"],
    ["Tδ utilisé", "=B17-B6", "Tmax - Tmin"],
]
for r, row_data in enumerate(params, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws_p.cell(row=r, column=c, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.border = thin_border
for r in [4,5,6,7,8,9,10,11,14]:
    ws_p.cell(r, 2).font = blue_font
    ws_p.cell(r, 2).fill = param_fill
ws_p.column_dimensions['A'].width = 25
ws_p.column_dimensions['B'].width = 16
ws_p.column_dimensions['C'].width = 40

# DONNÉES
ws_d = wb.create_sheet("Données")
data_cols = ['matricule', 'hrs', 'brut', 'ifm', 'iccp', 'cotpat', 'reduction', 'marge']
data_headers = ['Matricule', 'Hrs', 'Brut réel', 'IFM', 'ICCP', 'Cot.Pat.', 'Réduction', 'Marge']
for c, h in enumerate(data_headers, 1):
    cell = ws_d.cell(row=1, column=c, value=h)
    cell.font = bold_white
    cell.fill = header_fill
    cell.border = thin_border
for i, (_, row) in enumerate(df_clean.iterrows(), 2):
    for c_idx, key in enumerate(data_cols, 1):
        col_name = col_map.get(key)
        val = row[col_name] if col_name and col_name in row.index else 0
        if pd.isna(val): val = 0
        ws_d.cell(row=i, column=c_idx, value=val).border = thin_border

# SIMULATION
ws_s = wb.create_sheet("Simulation")
scenarios = [0, 10, 20, 30, 50, 100]
base_cols = 5
cols_per = 5
row2 = 3
for c, h in enumerate(['Matricule', 'Hrs', 'Brut', 'IFM', 'ICCP'], 1):
    cell = ws_s.cell(row2, c, h)
    cell.font = bold_white
    cell.fill = header_fill
    cell.border = thin_border

for s_idx, scen in enumerate(scenarios):
    sc = base_cols + 1 + s_idx * cols_per
    ws_s.cell(2, sc, f"CET {scen}%/{scen}%").font = Font(name='Arial', bold=True, size=11)
    ws_s.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=sc + cols_per - 1)
    for j, sh in enumerate(['Nv Brut', 'Ratio', 'C', 'Réd.', 'Gain']):
        cell = ws_s.cell(row2, sc + j, sh)
        cell.font = bold_white
        cell.fill = header_fill
        cell.border = thin_border

data_start = row2 + 1
for i in range(n_rows):
    r = data_start + i
    d_r = i + 2
    for c, d_col in enumerate(['A', 'B', 'C', 'D', 'E'], 1):
        ws_s.cell(r, c, f"=Données!{d_col}{d_r}").font = green_font
        ws_s.cell(r, c).border = thin_border
    for s_idx, scen in enumerate(scenarios):
        sc = base_cols + 1 + s_idx * cols_per
        pct = scen / 100.0
        nb = get_column_letter(sc)
        ws_s.cell(r, sc, f"=C{r}-D{r}*{pct}-E{r}*{pct}").border = thin_border
        ws_s.cell(r, sc).number_format = '#,##0.00 €'
        rat = get_column_letter(sc+1)
        ws_s.cell(r, sc+1, f"=IF({nb}{r}=0,0,3*Paramètres!$B$4*B{r}/{nb}{r})").border = thin_border
        ws_s.cell(r, sc+1).number_format = '0.0000'
        cc = get_column_letter(sc+2)
        ws_s.cell(r, sc+2, f'=IF({rat}{r}<=1,0,MIN(Paramètres!$B$6+Paramètres!$B$18*POWER(0.5*({rat}{r}-1),Paramètres!$B$11),Paramètres!$B$17))').border = thin_border
        ws_s.cell(r, sc+2).number_format = '0.000000'
        ws_s.cell(r, sc+3, f"={cc}{r}*{nb}{r}").border = thin_border
        ws_s.cell(r, sc+3).number_format = '#,##0.00 €'
        red0 = get_column_letter(base_cols + 1 + 3)
        red_c = get_column_letter(sc+3)
        ws_s.cell(r, sc+4, f"={red_c}{r}-{red0}{r}").border = thin_border
        ws_s.cell(r, sc+4).number_format = '#,##0.00 €'

tot_r = data_start + n_rows
ws_s.cell(tot_r, 1, "TOTAUX").font = Font(name='Arial', bold=True, size=11)
for s_idx in range(len(scenarios)):
    sc = base_cols + 1 + s_idx * cols_per
    for off in [0, 3, 4]:
        col = sc + off
        cl = get_column_letter(col)
        ws_s.cell(tot_r, col, f"=SUM({cl}{data_start}:{cl}{tot_r-1})").number_format = '#,##0.00 €'
        ws_s.cell(tot_r, col).font = Font(name='Arial', bold=True)

ws_s.freeze_panes = 'F4'
wb.save(OUTPUT_FILE)
print(f"Done. {n_rows} employees, {len(scenarios)} scenarios → {OUTPUT_FILE}")
